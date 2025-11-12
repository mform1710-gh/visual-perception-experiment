# save this as optimize_housing_sim.py
import math, random, time, csv
from pathlib import Path
from openpyxl import load_workbook, Workbook

# --------- CONFIG ----------
INPUT_XLSX = "20251112.xlsx"   # put uploaded file in same folder or use full path
OUT_XLSX = "optimized_with_housing_20251112.xlsx"
OUT_CSV  = "best_config_with_housing.csv"

# User-specified bounds (from you)
MAX_B = 50
MAX_C = 50
P5_MIN = 0.03
P5_MAX = 0.07

# Time budget (seconds) to avoid very long runs
TIME_BUDGET = 120.0

# Random + hillclimb settings (adjust to tradeoff speed vs quality)
NUM_RANDOM = 400
LOCAL_ITERS = 1000

# --------------------------

# Load workbook and read parameters + sheet
wb = load_workbook(INPUT_XLSX, data_only=True)
ws = wb.active

# Read param map O2:O10 -> P2:P10 if present (heuristic keys)
param_map = {}
for r in range(2,11):
    key = ws[f"O{r}"].value
    val = ws[f"P{r}"].value
    if key:
        param_map[str(key).strip()] = val

# try to infer main parameters (fallback to defaults)
PMT   = float(param_map.get("연간 예치금", param_map.get("연간예치금", 24000000)) or 24000000)
INIT  = float(param_map.get("초기예치금", param_map.get("조건2 초기예치", 300000000)) or 300000000)
HOUSE = float(param_map.get("주택가격", param_map.get("주택 1채 가격", 800000000)) or 800000000)
R_RETURN = float(param_map.get("예치금 반환", 0.03) or 0.03)
TERM = int(param_map.get("총 기간", 20) or 20)
CHURN = float(param_map.get("중도해지율", 0.10) or 0.10)
REFUND_RATIO = float(param_map.get("예비비비율", 0.70) or 0.70)

# read P16 initial cash explicitly (if present)
try:
    initial_cash_val = ws["P16"].value
    INITIAL_CASH = float(initial_cash_val) if initial_cash_val is not None else 0.0
except:
    INITIAL_CASH = 0.0

# read purchase allowed flags from I2:I40  -> map year index 1..39 to True/False
purchase_allowed = {}
for r in range(2,41):
    purchase_allowed[r-1] = bool(ws[f"I{r}"].value)

YEARS = 39  # rows 2..40

# --- Simulator implementing cohort tracking, churn, reserve, purchases, maturity payouts
def simulate(B_list, C_list, R_fund):
    balance = float(INITIAL_CASH)
    cohorts = []  # list of dicts {type, year, count, cum_paid, active}
    houses_bought = 0
    required_houses = int(sum(C_list))
    balances = []
    # compute reserve total and spread it evenly
    total_expected = sum(B_list[:TERM])*PMT + sum(C_list[:TERM])*(PMT*TERM + INIT)
    reserve_total = 0.2 * total_expected * REFUND_RATIO
    annual_reserve = reserve_total / YEARS if YEARS>0 else 0.0

    for t in range(YEARS):
        # add cohorts joining this year
        if B_list[t] > 0:
            cohorts.append({'type':'B','year':t,'count':int(B_list[t]),'cum_paid':0.0,'active':True})
        if C_list[t] > 0:
            cohorts.append({'type':'C','year':t,'count':int(C_list[t]),'cum_paid':0.0,'active':True,'init_paid':True})
        # inflows: PMT from each active person + INIT once for new Cs
        inflow = 0.0
        for c in cohorts:
            if not c['active']: continue
            inflow += c['count'] * PMT
            c['cum_paid'] += PMT
        for c in cohorts:
            if c['type']=='C' and c['year']==t and c.get('init_paid', False):
                inflow += c['count'] * INIT
                c['cum_paid'] += INIT
                c['init_paid'] = False
        balance += inflow
        balance *= (1 + R_fund)
        # purchase if allowed and needed
        if purchase_allowed.get(t+1, False):
            remaining = required_houses - houses_bought
            if remaining > 0 and balance >= HOUSE:
                can_buy = int(balance // HOUSE)
                to_buy = min(can_buy, remaining)
                balance -= to_buy * HOUSE
                houses_bought += to_buy
        # churn refunds
        refund_total = 0.0
        for c in cohorts:
            if not c['active']: continue
            exits = c['count'] * CHURN
            if exits > 0:
                refund_per = REFUND_RATIO * c['cum_paid']
                refund_total += exits * refund_per
                c['count'] -= exits
        balance -= refund_total
        # reserve deduction spread
        balance -= annual_reserve
        # maturity payouts
        maturity = 0.0
        for c in cohorts:
            if not c['active']: continue
            if (t - c['year']) + 1 >= TERM:
                rf = R_RETURN
                if rf == 0:
                    return_pmts = PMT * TERM
                else:
                    return_pmts = PMT * (((1+rf)**TERM - 1) / rf)
                if c['type']=='C':
                    per_person = return_pmts + INIT * ((1+rf)**TERM)
                else:
                    per_person = return_pmts
                maturity += c['count'] * per_person
                c['active'] = False
                c['count'] = 0
        balance -= maturity
        balances.append(balance)
        # early stop if extremely negative
        if balance < -1e14:
            break
    return balances, balance, houses_bought

# --- Search: random + local hillclimb
start_time = time.time()
best_score = -1e30
best_candidate = None

def rand_cand():
    B = [random.randint(0, MAX_B) for _ in range(YEARS)]
    C = [random.randint(0, MAX_C) for _ in range(YEARS)]
    p5 = random.uniform(P5_MIN, P5_MAX)
    return {'B':B,'C':C,'p5':p5}

for i in range(NUM_RANDOM):
    cand = rand_cand()
    balances, final, houses = simulate(cand['B'], cand['C'], cand['p5'])
    min_bal = min(balances) if balances else -1e30
    score = final - 1000.0 * max(0.0, -min_bal)
    if score > best_score:
        best_score = score
        best_candidate = {'B':cand['B'][:], 'C':cand['C'][:], 'p5':cand['p5'], 'final':final, 'min_bal':min_bal, 'houses':houses}
    if time.time() - start_time > TIME_BUDGET*0.6:
        break

# local hillclimb from best candidate
if best_candidate:
    cand = {'B':best_candidate['B'][:],'C':best_candidate['C'][:],'p5':best_candidate['p5']}
    cur_score = best_score
    for it in range(LOCAL_ITERS):
        if time.time() - start_time > TIME_BUDGET:
            break
        new = {'B':cand['B'][:],'C':cand['C'][:],'p5':cand['p5']}
        # small mutations
        for _ in range(3):
            y = random.randrange(YEARS)
            if random.random() < 0.5:
                new['B'][y] = max(0, min(MAX_B, new['B'][y] + random.randint(-2,2)))
            else:
                new['C'][y] = max(0, min(MAX_C, new['C'][y] + random.randint(-2,2)))
        if random.random() < 0.3:
            new['p5'] = max(P5_MIN, min(P5_MAX, new['p5'] + random.uniform(-0.001,0.001)))
        balances_n, final_n, houses_n = simulate(new['B'], new['C'], new['p5'])
        min_bal_n = min(balances_n) if balances_n else -1e30
        score_n = final_n - 1000.0 * max(0.0, -min_bal_n)
        if score_n > cur_score:
            cand = new
            cur_score = score_n
            if score_n > best_score:
                best_score = score_n
                best_candidate = {'B':cand['B'][:],'C':cand['C'][:],'p5':cand['p5'],'final':final_n,'min_bal':min_bal_n,'houses':houses_n}

# Save best solution back into Excel copy
if best_candidate:
    wb_out = load_workbook(INPUT_XLSX)
    ws_out = wb_out.active
    try:
        ws_out["P5"] = best_candidate['p5']
    except:
        pass
    for idx, r in enumerate(range(2, 41)):
        ws_out[f"B{r}"] = int(best_candidate['B'][idx])
        ws_out[f"C{r}"] = int(best_candidate['C'][idx])
    wb_out.save(OUT_XLSX)
    # CSV
    with open(OUT_CSV, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Year","B","C"])
        for idx in range(YEARS):
            w.writerow([idx+1, best_candidate['B'][idx], best_candidate['C'][idx]])

# Print summary
print("BEST SCORE:", best_score)
if best_candidate:
    print("Final balance (approx):", best_candidate['final'])
    print("Min balance:", best_candidate['min_bal'])
    print("Best P5:", best_candidate['p5'])
    print("Houses bought (approx):", best_candidate['houses'])
    print("Output workbook:", OUT_XLSX)
    print("Output csv:", OUT_CSV)
else:
    print("No feasible candidate found.")
